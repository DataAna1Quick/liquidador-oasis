#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════════════════════
 PROCESADOR GEOVICTORIA → OASIS — Quick Help SAS
═══════════════════════════════════════════════════════════════════════════════
 Versión: 6.0 (definitiva consolidada — Junio 2026)
 
 Convierte el reporte de asistencia de Geovictoria al formato de carga Oasis,
 aplicando las reglas operativas y contractuales de Quick Help.

 USO:
   python procesador_geovictoria_oasis.py <archivo_geovictoria.xlsx>
   python procesador_geovictoria_oasis.py input.xlsx -o salida.xlsx

 REGLAS APLICADAS:
   1. HEA (aprobadas) reemplaza a HEC: "lo aprobado es ley"
   2. Dominical NO se paga (cubierto por contrato dom-a-dom con compensatorio)
      EXCEPTO el componente nocturno (35%) que sí se paga aparte (RN)
   3. Festivo SÍ se paga completo (RDF 75%, RNF 125%)
   4. HE en dominical → HED/HEN ordinaria
   5. HE en festivo → HEFD/HEFN completa con plus festivo
   6. DOD/DON/DHD/DHN NUNCA se pagan
   7. JUSTIFICAR MOTIVO siempre vacío
   8. CODIGO CENTRO DE COSTO = parte numérica del Grupo (ej. "2003 - 1003")
   9. Solo filas con pago > 0 (no totalizar, no incluir días sin pago)

 SALIDA:
   - Hoja única "Oasis_Liquidacion" con 24 columnas
   - Fila 1: headers, Fila 2: factores, Filas 3+: datos
   - Filas con ajuste HEA > HEC marcadas en verde
   - Auto-filter, freeze panes G3
   - Reporte por consola con totales, conciliación y alertas
═══════════════════════════════════════════════════════════════════════════════
"""

import pandas as pd
import re
import sys
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════════════════

FESTIVOS_2026 = {
    '01-01': 'Año Nuevo',
    '06-01': 'Reyes Magos',
    '23-03': 'San José',
    '02-04': 'Jueves Santo',
    '03-04': 'Viernes Santo',
    '01-05': 'Día del Trabajo',
    '18-05': 'Ascensión',
    '08-06': 'Corpus Christi',
    '15-06': 'Sagrado Corazón',
    '29-06': 'San Pedro y San Pablo',
    '20-07': 'Independencia',
    '07-08': 'Batalla de Boyacá',
    '17-08': 'Asunción',
    '12-10': 'Día de la Raza',
    '02-11': 'Todos los Santos',
    '16-11': 'Independencia de Cartagena',
    '08-12': 'Inmaculada Concepción',
    '25-12': 'Navidad',
}

GEOVICTORIA_COLS = [
    'Apellidos','Nombres','Identificador','Grupo','Fecha','Permiso','Turno',
    'Entro1','Atraso1','Salio1','Adelanto1','Entro2','Atraso2','Salio2','Adelanto2',
    'HEA','HEC','HT','HNT','HP',
    'HEDO','HEDO_Dec','HENO','HENO_Dec','HEDD','HEDD_Dec','HEND','HEND_Dec',
    'HEDF','HEDF_Dec','HENF','HENF_Dec',
    'RNO','RNO_Dec','RDDC','RDDC_Dec','RDNC','RDNC_Dec','RDD','RDD_Dec',
    'RND','RND_Dec','RDFC','RDFC_Dec','RNFC','RNFC_Dec','RDF','RDF_Dec',
    'RNF','RNF_Dec','Cargo'
]

OASIS_COLS = [
    'BDT','CENTRO DE TRABAJO','CIUDAD','IDENTIFICACION','COLABORADOR','CARGO',
    'RN','RDF','RNF','DOD','DON','DHD','DHN','HED','HEN','HEFD','HEFN',
    'FECHA','HORA INICIO','HORA FINAL','JUSTIFICAR MOTIVO',
    'CODIGO CENTRO DE COSTO','NOMBRE CENTRO DE COSTO','COBRAR CLIENTE'
]

FACTORES = {
    'RN': 0.35, 'RDF': 0.75, 'RNF': 1.25,
    'DOD': 1.75, 'DON': 2.50, 'DHD': 1.75, 'DHN': 2.50,
    'HED': 1.25, 'HEN': 1.75, 'HEFD': 2.00, 'HEFN': 2.50,
}

CIUDADES = ['BOGOTA','MEDELLIN','MEDELLÍN','CALI','BARRANQUILLA','BARRANQUILA',
            'FUNZA','SANTANDER','EJE CAFETERO','COSTA','CARTAGENA','BUCARAMANGA']

# Estilos visuales
STYLE_HEADER_FILL = PatternFill('solid', start_color='1F4E78')
STYLE_HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
STYLE_FACTOR_FILL = PatternFill('solid', start_color='D9E1F2')
STYLE_FACTOR_FONT = Font(name='Arial', bold=True, color='1F4E78', size=10, italic=True)
STYLE_NORMAL_FONT = Font(name='Arial', size=9)
STYLE_AJUSTE_FILL = PatternFill('solid', start_color='E2EFDA')
STYLE_AJUSTE_FONT = Font(name='Arial', size=9, color='375623')
STYLE_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'))

COL_WIDTHS = [6,40,15,14,32,32,7,7,7,6,6,6,6,7,7,7,7,12,11,11,18,20,40,12]

# ═══════════════════════════════════════════════════════════════════════════════
# UTILIDADES DE CONVERSIÓN
# ═══════════════════════════════════════════════════════════════════════════════

def hms_to_dec(v):
    """Convierte 'HH:MM' o 'HH:MM:SS' a horas decimales."""
    if pd.isna(v): return 0.0
    s = str(v).strip()
    if ':' in s:
        parts = s.split(':')
        return int(parts[0]) + int(parts[1])/60 + (int(parts[2])/3600 if len(parts) > 2 else 0)
    try: return float(s.replace(',', '.'))
    except: return 0.0

def to_float(v):
    """Convierte a float manejando coma decimal colombiana."""
    if pd.isna(v) or v == '': return 0.0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip().replace(',', '.'))
    except: return 0.0

def parse_fecha(s):
    """'Vie 01-05-2026' → '01/05/2026'"""
    m = re.match(r'\w+\s+(\d{2})-(\d{2})-(\d{4})', s)
    return f"{m.group(1)}/{m.group(2)}/{m.group(3)}" if m else s

def primera_marca(r, cols=('Entro1', 'Entro2')):
    """Primera marcación válida del día."""
    for c in cols:
        v = str(r[c])
        if re.match(r'\d{2}:\d{2}', v): return v
    return ''

def ultima_marca(r, cols=('Salio2', 'Salio1')):
    """Última marcación válida del día."""
    for c in cols:
        v = str(r[c])
        if re.match(r'\d{2}:\d{2}', v): return v
    return ''

def extraer_ciudad(grupo):
    """Detecta la ciudad dentro del nombre del Grupo."""
    g = grupo.upper()
    for c in CIUDADES:
        if c in g:
            return c.replace('Í', 'I').replace('BARRANQUILA', 'BARRANQUILLA')
    return ''

def extraer_codigo_cco(grupo):
    """
    Extrae código numérico del Grupo.
    Ejemplo: '2003 - 1003 POLLOS EL BUCANERO' → '2003 - 1003'
    """
    m = re.match(r'^\s*(\d+)\s*-\s*(\d+)', grupo)
    return f"{m.group(1)} - {m.group(2)}" if m else ''

def es_festivo(fecha_str):
    """Detecta si una fecha es festivo nacional colombiano (NO domingos)."""
    m = re.search(r'(\d{2})-(\d{2})-(\d{4})', fecha_str)
    if not m: return False
    dia, mes, _ = m.groups()
    return f"{dia}-{mes}" in FESTIVOS_2026

def turno_es_nocturno(turno):
    """Detecta si un turno cruza la franja nocturna (19:00-06:00, Ley 2466/2025)."""
    if not turno or turno in ('Descanso', 'No Planificado', 'Sin turno', ''):
        return False
    matches = re.findall(r'(\d{2}):(\d{2})', turno)
    if len(matches) < 2: return False
    h_ini = int(matches[0][0])
    h_fin = int(matches[1][0])
    return h_ini >= 19 or h_fin < h_ini or h_fin >= 22

# ═══════════════════════════════════════════════════════════════════════════════
# LÓGICA DE MAPEO (REGLAS DEFINITIVAS)
# ═══════════════════════════════════════════════════════════════════════════════

def mapear_pago(r):
    """
    Mapea una fila de Geovictoria a los conceptos de Oasis.
    
    Recargos:
    - RN  ← RNO + RND + RDNC (incluye componente nocturno de dominicales)
    - RDF ← RDF + RDFC (solo festivos diurnos; RDD/RDDC NO se pagan)
    - RNF ← RNF + RNFC (solo festivos nocturnos)
    
    Horas Extras (basadas en HEA aprobadas):
    - HED  ← HEDO + HEDD (ordinaria + dominical como ordinaria)
    - HEN  ← HENO + HEND
    - HEFD ← HEDF (festivo diurno completo)
    - HEFN ← HENF (festivo nocturno completo)
    """
    hea = r['HEA_dec']
    hec = r['HEC_dec']
    pago = {'HED': 0.0, 'HEN': 0.0, 'HEFD': 0.0, 'HEFN': 0.0}

    if hea > 0:
        if hec > 0:
            total_desglose = (r['HEDO_Dec'] + r['HENO_Dec'] + r['HEDD_Dec'] +
                              r['HEND_Dec'] + r['HEDF_Dec'] + r['HENF_Dec'])
            if total_desglose > 0:
                pago['HED']  = hea * ((r['HEDO_Dec'] + r['HEDD_Dec']) / total_desglose)
                pago['HEN']  = hea * ((r['HENO_Dec'] + r['HEND_Dec']) / total_desglose)
                pago['HEFD'] = hea * (r['HEDF_Dec'] / total_desglose)
                pago['HEFN'] = hea * (r['HENF_Dec'] / total_desglose)
            else:
                _asignar_por_turno(pago, hea, r)
        else:
            _asignar_por_turno(pago, hea, r)

    return {
        'RN':   r['RNO_Dec'] + r['RND_Dec'] + r['RDNC_Dec'],
        'RDF':  r['RDF_Dec'] + r['RDFC_Dec'],
        'RNF':  r['RNF_Dec'] + r['RNFC_Dec'],
        'HED':  pago['HED'],
        'HEN':  pago['HEN'],
        'HEFD': pago['HEFD'],
        'HEFN': pago['HEFN'],
    }

def _asignar_por_turno(pago, hea, r):
    """
    Cuando HEC=0 (aprobada pero no marcada), distribuir HEA según turno y fecha.
    NOTA: dominicales NO son festivos en esta lógica (van a HED ordinaria).
    """
    fest = es_festivo(r['Fecha'])
    noct = turno_es_nocturno(r['Turno'])
    if fest and noct:   pago['HEFN'] = hea
    elif fest:          pago['HEFD'] = hea
    elif noct:          pago['HEN']  = hea
    else:               pago['HED']  = hea

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESAMIENTO PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════

def cargar_geovictoria(path):
    """Carga y normaliza el archivo de Geovictoria."""
    df = pd.read_excel(path, header=1)
    
    if len(df.columns) != len(GEOVICTORIA_COLS):
        print(f"⚠️  ADVERTENCIA: archivo tiene {len(df.columns)} columnas, "
              f"esperadas {len(GEOVICTORIA_COLS)}")
    df.columns = GEOVICTORIA_COLS

    # Conversiones HH:MM y decimales
    df['HEA_dec'] = df['HEA'].apply(hms_to_dec)
    df['HEC_dec'] = df['HEC'].apply(hms_to_dec)
    for c in [c for c in GEOVICTORIA_COLS if c.endswith('_Dec')]:
        df[c] = df[c].apply(to_float)

    # Strings limpios
    for c in ['Apellidos', 'Nombres', 'Grupo', 'Fecha', 'Permiso', 'Turno',
              'Cargo', 'Entro1', 'Salio1', 'Entro2', 'Salio2', 'HT']:
        df[c] = df[c].apply(lambda v: '' if pd.isna(v) else str(v).strip())

    return df

def construir_filas_oasis(df):
    """Construye la lista de filas para el archivo Oasis."""
    filas = []
    for _, r in df.iterrows():
        p = mapear_pago(r)
        if sum(p.values()) <= 0:
            continue

        filas.append({
            'BDT': '',
            'CENTRO DE TRABAJO': r['Grupo'],
            'CIUDAD': extraer_ciudad(r['Grupo']),
            'IDENTIFICACION': r['Identificador'],
            'COLABORADOR': f"{r['Apellidos']} {r['Nombres']}".strip(),
            'CARGO': r['Cargo'],
            'RN':   round(p['RN'], 4)   if p['RN']   > 0 else None,
            'RDF':  round(p['RDF'], 4)  if p['RDF']  > 0 else None,
            'RNF':  round(p['RNF'], 4)  if p['RNF']  > 0 else None,
            'DOD': None, 'DON': None, 'DHD': None, 'DHN': None,
            'HED':  round(p['HED'], 4)  if p['HED']  > 0 else None,
            'HEN':  round(p['HEN'], 4)  if p['HEN']  > 0 else None,
            'HEFD': round(p['HEFD'], 4) if p['HEFD'] > 0 else None,
            'HEFN': round(p['HEFN'], 4) if p['HEFN'] > 0 else None,
            'FECHA': parse_fecha(r['Fecha']),
            'HORA INICIO': primera_marca(r),
            'HORA FINAL':  ultima_marca(r),
            'JUSTIFICAR MOTIVO': '',
            'CODIGO CENTRO DE COSTO': extraer_codigo_cco(r['Grupo']),
            'NOMBRE CENTRO DE COSTO': r['Grupo'],
            'COBRAR CLIENTE': '',
            '_es_ajuste': r['HEA_dec'] > r['HEC_dec'] + 0.01,
        })

    odf = pd.DataFrame(filas)
    if len(odf) > 0:
        odf['_fs'] = pd.to_datetime(odf['FECHA'], format='%d/%m/%Y', errors='coerce')
        odf = odf.sort_values(['CENTRO DE TRABAJO', 'COLABORADOR', '_fs'])\
                 .drop(columns=['_fs']).reset_index(drop=True)
    return odf

def escribir_excel(odf, output_path):
    """Genera el archivo Excel con formato Oasis exacto."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('Oasis_Liquidacion')

    # Fila 1: Headers
    for i, h in enumerate(OASIS_COLS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = STYLE_HEADER_FILL
        c.font = STYLE_HEADER_FONT
        c.border = STYLE_BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Fila 2: Factores
    for i, h in enumerate(OASIS_COLS, 1):
        c = ws.cell(row=2, column=i, value=FACTORES.get(h))
        c.font = STYLE_FACTOR_FONT
        c.fill = STYLE_FACTOR_FILL
        c.border = STYLE_BORDER
        c.alignment = Alignment(horizontal='center')
        if h in FACTORES:
            c.number_format = '0.00'

    # Datos
    hour_cols = {7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17}
    for ri, rec in enumerate(odf.to_dict('records'), start=3):
        es_aj = rec.get('_es_ajuste', False)
        for ci, name in enumerate(OASIS_COLS, start=1):
            c = ws.cell(row=ri, column=ci, value=rec.get(name))
            if es_aj:
                c.font = STYLE_AJUSTE_FONT
                c.fill = STYLE_AJUSTE_FILL
            else:
                c.font = STYLE_NORMAL_FONT
            c.border = STYLE_BORDER
            if ci in hour_cols and rec.get(name) is not None:
                c.number_format = '0.00'

    # Layout
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'G3'
    if len(odf) > 0:
        ws.auto_filter.ref = f'A1:{get_column_letter(len(OASIS_COLS))}{2+len(odf)}'

    wb.save(output_path)

def detectar_periodo(df):
    """Detecta el rango de fechas del archivo."""
    fechas = df['Fecha'].dropna().unique()
    if len(fechas) == 0: return ('?', '?')
    parseadas = sorted([parse_fecha(f) for f in fechas])
    return (parseadas[0], parseadas[-1])

def reporte(df, odf, output_path):
    """Imprime el reporte resumen estándar."""
    periodo = detectar_periodo(df)
    print(f"\n{'═'*72}")
    print(f"  ✅ ARCHIVO GENERADO: {output_path}")
    print(f"{'═'*72}\n")

    print(f"📊 RESUMEN")
    print(f"  • Periodo:               {periodo[0]} a {periodo[1]}")
    print(f"  • Filas Oasis:           {len(odf)}")
    print(f"  • Empleados con pago:    {odf['IDENTIFICACION'].nunique() if len(odf) > 0 else 0} "
          f"de {df['Identificador'].nunique()} totales")
    print(f"  • Ajustes HEA>HEC:       {int(odf['_es_ajuste'].sum()) if len(odf) > 0 else 0} filas (verdes)")

    # Totales por concepto
    totales = {c: pd.to_numeric(odf[c], errors='coerce').fillna(0).sum()
               for c in ['RN','RDF','RNF','HED','HEN','HEFD','HEFN']}

    print(f"\n💰 TOTALES POR CONCEPTO")
    print(f"  {'Concepto':6s} {'Horas':>10s}  {'Factor':>6s}  {'Equivalente':>12s}")
    print(f"  {'─'*44}")
    total_h = 0.0
    total_eq = 0.0
    for k, v in totales.items():
        if v > 0:
            eq = v * FACTORES[k]
            total_h += v
            total_eq += eq
            print(f"  {k:6s} {v:10.2f}  {FACTORES[k]:>6.2f}  {eq:12.2f}")
    print(f"  {'─'*44}")
    print(f"  {'TOTAL':6s} {total_h:10.2f}        {total_eq:12.2f}")

    # Conciliación
    hea_total = df['HEA_dec'].sum()
    he_oasis = totales['HED'] + totales['HEN'] + totales['HEFD'] + totales['HEFN']
    print(f"\n🔍 CONCILIACIÓN")
    print(f"  HEA aprobadas (Geo):  {hea_total:.2f} h")
    print(f"  HE en Oasis (output): {he_oasis:.2f} h")
    diff = abs(hea_total - he_oasis)
    print(f"  Diferencia:           {diff:.4f} h", 
          "✅ OK" if diff < 0.01 else "⚠️ revisar")

    # Alertas
    print(f"\n⚠️  ALERTAS")
    
    # Caso 3: HEC > 0 sin HEA
    caso3 = df[(df['HEA_dec'] == 0) & (df['HEC_dec'] > 0)]
    if len(caso3) > 0:
        print(f"  • Caso 3 (HEC>0 sin HEA): {len(caso3)} filas, {caso3['HEC_dec'].sum():.2f}h")
        for _, r in caso3.head(10).iterrows():
            print(f"    - {r['Apellidos']} {r['Nombres']} ({r['Fecha']}): {r['HEC_dec']:.2f}h")
        if len(caso3) > 10:
            print(f"    ... y {len(caso3)-10} más")
    else:
        print(f"  • Caso 3 (HEC>0 sin HEA): 0 filas ✅")

    # HE > 2h diarias
    df['HE_dia'] = (df['HEDO_Dec'] + df['HENO_Dec'] + df['HEDD_Dec'] + 
                    df['HEND_Dec'] + df['HEDF_Dec'] + df['HENF_Dec'])
    exceso = df[df['HE_dia'] > 2]
    if len(exceso) > 0:
        print(f"  • Días con HE > 2h (Art. 22 Ley 50/1990): {len(exceso)} casos")
    
    # HEA sin marcación
    df['marco'] = df['Entro1'].astype(str).apply(lambda x: bool(re.match(r'\d{2}:\d{2}', x)))
    sin_marca = df[(df['HEA_dec'] > 0) & (~df['marco'])]
    if len(sin_marca) > 0:
        print(f"  • HEA aprobada SIN marcación: {len(sin_marca)} filas, {sin_marca['HEA_dec'].sum():.2f}h")

    print()

# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Procesa archivo Geovictoria y genera formato Oasis (Quick Help SAS)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    parser.add_argument('input', type=str,
                        help='Archivo Geovictoria.xlsx de entrada')
    parser.add_argument('--output', '-o', type=str, default=None,
                        help='Archivo de salida (default: Oasis_Liquidacion_<input>.xlsx)')
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"❌ Error: archivo no encontrado: {input_path}")
        sys.exit(1)

    output_path = Path(args.output) if args.output else \
                  input_path.parent / f"Oasis_Liquidacion_{input_path.stem}.xlsx"

    print(f"\n📥 Cargando: {input_path}")
    df = cargar_geovictoria(input_path)
    print(f"   {len(df)} filas, {df['Identificador'].nunique()} empleados")

    print(f"⚙️  Procesando con reglas v6.0...")
    odf = construir_filas_oasis(df)

    print(f"💾 Escribiendo: {output_path}")
    escribir_excel(odf, output_path)

    reporte(df, odf, output_path)

if __name__ == '__main__':
    main()
